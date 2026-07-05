"""
Inspect the _v2 Excel file to understand its structure and lat/long status.
"""
import openpyxl

XLSX = r'e:\Cybersecurity\Apps\TEAAPP\ext\Registered Tea Factories All Island_v2.xlsx'
wb = openpyxl.load_workbook(XLSX, data_only=True)

print(f"Sheets: {wb.sheetnames}")
print()

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"=== Sheet: {sname} | rows={ws.max_row} cols={ws.max_column} ===")
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"Headers: {hdrs}")
    for i in range(2, min(4, ws.max_row + 1)):
        row = [ws.cell(i, c).value for c in range(1, ws.max_column + 1)]
        print(f"  Row {i}: {row}")
    print()

# Check for lat/long columns in each sheet
for sname in wb.sheetnames:
    ws = wb[sname]
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    lat_col = None
    lon_col = None
    for ci, h in enumerate(hdrs, 1):
        if h and ('lat' in str(h).lower() or 'latitude' in str(h).lower()):
            lat_col = ci
        if h and ('lon' in str(h).lower() or 'long' in str(h).lower()):
            lon_col = ci
    
    if lat_col or lon_col:
        with_ll = 0
        without_ll = 0
        for i in range(2, ws.max_row + 1):
            lat = ws.cell(i, lat_col).value if lat_col else None
            lon = ws.cell(i, lon_col).value if lon_col else None
            if lat is not None and lon is not None:
                with_ll += 1
            else:
                without_ll += 1
        print(f"Sheet '{sname}' lat(col {lat_col})/lon(col {lon_col}): WITH={with_ll}, WITHOUT={without_ll}")
    else:
        print(f"Sheet '{sname}': No lat/lon columns found")
    print()

# Find FacNo column across sheets and check join keys
for sname in wb.sheetnames:
    ws = wb[sname]
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    facno_col = None
    for ci, h in enumerate(hdrs, 1):
        if h and 'fac' in str(h).lower() and ('no' in str(h).lower() or 'num' in str(h).lower() or h.lower() == 'facno'):
            facno_col = ci
            break
    if facno_col:
        facnos = set(ws.cell(i, facno_col).value for i in range(2, ws.max_row + 1) if ws.cell(i, facno_col).value)
        print(f"Sheet '{sname}': FacNo col={facno_col} ({hdrs[facno_col-1]}), unique facnos={len(facnos)}")
        print(f"  Sample: {list(facnos)[:5]}")
    else:
        print(f"Sheet '{sname}': No FacNo column found")
    print()
