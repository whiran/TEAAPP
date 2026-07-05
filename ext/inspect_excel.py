
import openpyxl, json

XLSX = r'e:\Cybersecurity\Apps\TEAAPP\ext\Registered Tea Factories All Island.xlsx'
wb = openpyxl.load_workbook(XLSX, data_only=True)

output = []

for sname in wb.sheetnames:
    ws = wb[sname]
    output.append(f"\n=== Sheet: {sname} | rows={ws.max_row} cols={ws.max_column} ===")
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    output.append(f"Headers: {hdrs}")
    for i in range(2, min(6, ws.max_row + 1)):
        row = [ws.cell(i, c).value for c in range(1, ws.max_column + 1)]
        output.append(f"  Row {i}: {row}")

# Specifically check Sheet4 lat/long columns
ws4 = wb['Sheet4']
with_ll = 0
without_ll = 0
for i in range(2, ws4.max_row + 1):
    lat = ws4.cell(i, 9).value
    lng = ws4.cell(i, 10).value
    if lat is not None and lng is not None:
        with_ll += 1
    else:
        without_ll += 1
output.append(f"\nSheet4 lat/long: WITH={with_ll}, WITHOUT={without_ll}")

# Check Sheet1 Facno vs Sheet4 FacNo for join
ws1 = wb['Sheet1']
s1_facnos = set(ws1.cell(i, 1).value for i in range(2, ws1.max_row + 1) if ws1.cell(i, 1).value)
s4_facnos = set(ws4.cell(i, 1).value for i in range(2, ws4.max_row + 1) if ws4.cell(i, 1).value)
output.append(f"\nSheet1 Facnos: {len(s1_facnos)} | Sheet4 FacNos: {len(s4_facnos)}")
output.append(f"Common: {len(s1_facnos & s4_facnos)}")
output.append(f"Sample S1 facnos: {list(s1_facnos)[:5]}")
output.append(f"Sample S4 facnos: {list(s4_facnos)[:5]}")

# Check Sheet2 and Sheet3 for lat/long
for sn in wb.sheetnames:
    ws = wb[sn]
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    output.append(f"\n{sn} full headers: {hdrs}")

result = "\n".join(str(o) for o in output)
with open(r'e:\Cybersecurity\Apps\TEAAPP\ext\excel_inspect_result.txt', 'w', encoding='utf-8') as f:
    f.write(result)

print("Done - saved to excel_inspect_result.txt")
print(result)
