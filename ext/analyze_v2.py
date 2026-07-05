"""
Analyze the _v2 Excel file in detail:
- Sheet1 has Latitude/Longitude (cols 21/22) - check which 2 are missing
- Sheet4 has Lant column with combined "lat, lon" strings
- Compare coordinates between sheets for discrepancies
"""
import openpyxl

XLSX = r'e:\Cybersecurity\Apps\TEAAPP\ext\Registered Tea Factories All Island_v2.xlsx'
wb = openpyxl.load_workbook(XLSX, data_only=True)

ws1 = wb['Sheet1']
ws4 = wb['Sheet4']

print("=" * 70)
print("SHEET1 ANALYSIS")
print("=" * 70)
hdrs1 = [ws1.cell(1, c).value for c in range(1, ws1.max_column + 1)]
print(f"All headers: {hdrs1}")
print(f"Lat col (idx 21): '{hdrs1[20]}', Lon col (idx 22): '{hdrs1[21]}'")
print(f"Url col (idx 23): '{hdrs1[22]}'")
print()

# Find rows missing coordinates in Sheet1
print("Rows MISSING coordinates in Sheet1:")
missing = []
for i in range(2, ws1.max_row + 1):
    lat = ws1.cell(i, 21).value
    lon = ws1.cell(i, 22).value
    if lat is None or lon is None:
        facno = ws1.cell(i, 1).value
        name = ws1.cell(i, 2).value
        address = ws1.cell(i, 10).value
        district = ws1.cell(i, 14).value
        missing.append({'row': i, 'facno': facno, 'name': name, 'address': address, 'district': district, 'lat': lat, 'lon': lon})
        print(f"  Row {i}: {facno} | {name} | {address} | {district} | lat={lat} lon={lon}")

print(f"\nTotal missing: {len(missing)}")
print()

print("=" * 70)
print("SHEET4 ANALYSIS")
print("=" * 70)
hdrs4 = [ws4.cell(1, c).value for c in range(1, ws4.max_column + 1)]
print(f"All headers: {hdrs4}")
print()

# Check Sheet4 Lant column format + Long column
print("Sheet4 coordinate data (first 10 rows):")
for i in range(2, min(12, ws4.max_row + 1)):
    name = ws4.cell(i, 2).value
    lant = ws4.cell(i, 9).value  # Lant column
    long = ws4.cell(i, 10).value  # Long column
    print(f"  Row {i}: {name} | Lant='{lant}' | Long='{long}'")

# Count Sheet4 rows with data in Lant
with_lant = 0
without_lant = 0
parseable = 0
for i in range(2, ws4.max_row + 1):
    lant = ws4.cell(i, 9).value
    if lant is not None:
        with_lant += 1
        # Try to parse as "lat, lon"
        try:
            parts = str(lant).split(',')
            if len(parts) == 2:
                float(parts[0].strip())
                float(parts[1].strip())
                parseable += 1
        except:
            pass
    else:
        without_lant += 1

print(f"\nSheet4 Lant col: with_data={with_lant}, without={without_lant}, parseable_as_lat_lon={parseable}")

print()
print("=" * 70)
print("COMPARISON: Sheet1 vs Sheet4 coordinates")
print("=" * 70)

# Compare coordinates between Sheet1 and Sheet4 (by row position, since Sheet4 has no FacNo)
print("\nFirst 5 rows comparison:")
for i in range(2, 7):
    s1_name = ws1.cell(i, 2).value
    s1_lat = ws1.cell(i, 21).value
    s1_lon = ws1.cell(i, 22).value
    s4_name = ws4.cell(i, 2).value
    s4_lant = ws4.cell(i, 9).value
    
    s4_lat, s4_lon = None, None
    if s4_lant:
        try:
            parts = str(s4_lant).split(',')
            s4_lat = float(parts[0].strip())
            s4_lon = float(parts[1].strip())
        except:
            pass
    
    lat_diff = abs(s1_lat - s4_lat) if s1_lat and s4_lat else 'N/A'
    lon_diff = abs(s1_lon - s4_lon) if s1_lon and s4_lon else 'N/A'
    
    print(f"Row {i}: S1={s1_name} lat={s1_lat} lon={s1_lon}")
    print(f"        S4={s4_name} lat={s4_lat} lon={s4_lon}")
    print(f"        Difference: lat_diff={lat_diff}, lon_diff={lon_diff}")
    print()
