"""
Script to:
1. Read Registered Tea Factories All Island_v2.xlsx (Sheet1)
2. Geocode the 2 missing factories (MF0160 Pussatenna, MF0572 Hoptan)
3. Write updated coordinates back to Sheet1
4. Also export a clean factories.json for the app

Missing factories:
- Row 86: MF0160 | Pussatenna | Pussatenne Estate,Panvilatenna,Gampola. | KANDY
- Row 289: MF0572 | Hoptan  | LunugalaBadulla | BADULLA
"""

import openpyxl
import requests
import time
import json
import os
from pathlib import Path

BASE_DIR = Path(__file__).parent.parent.parent
EXT_DIR = BASE_DIR / "ext"
FRONTEND_DATA = BASE_DIR / "frontend" / "public" / "data"

XLSX = EXT_DIR / "Registered Tea Factories All Island_v2.xlsx"
OUTPUT_XLSX = EXT_DIR / "Registered Tea Factories All Island_v2_geocoded.xlsx"
OUTPUT_JSON = FRONTEND_DATA / "factories.json"

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
USER_AGENT = "TeaFactoryGeocoder/2.0 (Ceylon Tea Intelligence Platform)"

def geocode(query, session):
    """Geocode a single address using Nominatim API."""
    params = {'q': query, 'format': 'json', 'limit': 1, 'countrycodes': 'lk'}
    headers = {'User-Agent': USER_AGENT}
    try:
        r = session.get(NOMINATIM_URL, params=params, headers=headers, timeout=10)
        r.raise_for_status()
        results = r.json()
        if results:
            return float(results[0]['lat']), float(results[0]['lon'])
    except Exception as e:
        print(f"  Error: {e}")
    return None, None


def main():
    print("=" * 60)
    print("Tea Factory Coordinate Update - _v2 file")
    print("=" * 60)
    
    wb = openpyxl.load_workbook(str(XLSX))
    ws1 = wb['Sheet1']
    
    # Get headers
    hdrs = [ws1.cell(1, c).value for c in range(1, ws1.max_column + 1)]
    print(f"Sheet1 columns: {len(hdrs)}")
    
    # Find column indices (1-indexed)
    facno_col = hdrs.index('Facno') + 1  # = 1
    name_col = hdrs.index('facName') + 1  # = 2
    address_col = hdrs.index('FacAddress') + 1  # = 10
    district_col = hdrs.index('AdminDistrict') + 1  # = 14
    lat_col = hdrs.index('Latitude') + 1  # = 21
    lon_col = hdrs.index('Longitude') + 1  # = 22
    
    print(f"  Lat col: {lat_col}, Lon col: {lon_col}")
    
    # Find and fix the 2 missing factories
    session = requests.Session()
    missing_fixed = 0
    
    for row_i in range(2, ws1.max_row + 1):
        lat = ws1.cell(row_i, lat_col).value
        lon = ws1.cell(row_i, lon_col).value
        
        if lat is None or lon is None:
            facno = ws1.cell(row_i, facno_col).value
            name = ws1.cell(row_i, name_col).value
            address = ws1.cell(row_i, address_col).value
            district = ws1.cell(row_i, district_col).value
            
            print(f"\nGeocoding Row {row_i}: {facno} | {name}")
            print(f"  Address: {address}")
            print(f"  District: {district}")
            
            # Try multiple queries
            queries = []
            if address:
                queries.append(f"{address.strip()}, {district.strip() if district else ''}, Sri Lanka")
                queries.append(f"{address.strip()}, Sri Lanka")
            if district:
                queries.append(f"{district.strip()}, Sri Lanka")
            
            found_lat, found_lon = None, None
            for q in queries:
                print(f"  Trying query: '{q}'")
                found_lat, found_lon = geocode(q, session)
                time.sleep(1.5)
                if found_lat is not None:
                    print(f"  SUCCESS: lat={found_lat}, lon={found_lon}")
                    break
                else:
                    print(f"  No result.")
            
            if found_lat is not None:
                ws1.cell(row_i, lat_col).value = found_lat
                ws1.cell(row_i, lon_col).value = found_lon
                missing_fixed += 1
            else:
                print(f"  FAILED: Could not geocode {name}")
    
    session.close()
    print(f"\nFixed {missing_fixed} missing coordinates.")
    
    # Save updated xlsx
    print(f"\nSaving updated xlsx to: {OUTPUT_XLSX}")
    wb.save(str(OUTPUT_XLSX))
    print("Saved!")
    
    # Export factories.json
    print(f"\nExporting factories.json...")
    factories = []
    
    # Re-read to get clean data
    wb2 = openpyxl.load_workbook(str(OUTPUT_XLSX), data_only=True)
    ws = wb2['Sheet1']
    hdrs2 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    
    # Map column names to indices
    col_map = {h: i+1 for i, h in enumerate(hdrs2) if h}
    
    no_coords = 0
    for row_i in range(2, ws.max_row + 1):
        lat = ws.cell(row_i, col_map['Latitude']).value
        lon = ws.cell(row_i, col_map['Longitude']).value
        
        if lat is None or lon is None:
            no_coords += 1
            continue
        
        def get(col_name):
            idx = col_map.get(col_name)
            if idx is None:
                return None
            v = ws.cell(row_i, idx).value
            if v is None:
                return None
            if isinstance(v, str):
                v = v.strip()
            return v
        
        factory = {
            "facno": get('Facno'),
            "name": get('facName'),
            "address": get('FacAddress'),
            "year": get('YearEstablished'),
            "method": get('Method'),
            "isActive": bool(get('IsActive')),
            "gnDiv": get('GramaNiladhariDivision'),
            "village": get('Village'),
            "elevHeight": get('ElevationHeight'),
            "greenTeaAvg": get('GreenTeaAvgQty'),
            "capacity": get('capacity'),
            "elevation": get('Elvation'),
            "subElevation": get('subelevation'),
            "district": get('AdminDistrict'),
            "dsDiv": get('DSDivision'),
            "atcReg": get('ATCReg'),
            "inspectorRegion": get('InspectorRegion'),
            "agroDiv": get('Agrodiv'),
            "managementType": get('ManagementType'),
            "lat": float(lat),
            "lng": float(lon),
        }
        factories.append(factory)
    
    print(f"  Total rows: {ws.max_row - 1}")
    print(f"  With coordinates: {len(factories)}")
    print(f"  Without coordinates (skipped): {no_coords}")
    
    # Ensure output directory exists
    FRONTEND_DATA.mkdir(parents=True, exist_ok=True)
    
    with open(str(OUTPUT_JSON), 'w', encoding='utf-8') as f:
        json.dump(factories, f, ensure_ascii=False, default=str)
    
    print(f"  Saved to: {OUTPUT_JSON}")
    print(f"\n{'=' * 60}")
    print("Done!")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
